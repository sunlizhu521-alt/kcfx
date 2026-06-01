from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, timezone
import json
import math


FILES = {
    "fact-inventory": Path(r"D:\BI文件\维度表刷新\物料收发汇总表_2026052411192401_115816.xlsx"),
    "dim-product": Path(r"D:\BI文件\维度表\2026年维度表\Dim-YL医疗器械商品分类-2026年整理版.xlsx"),
    "dim-warehouse": Path(r"D:\BI文件\维度表\2026年维度表\Dim-仓库_金蝶、旺店通、领星-2026年整理版.xlsx"),
    "dim-warehouse-material": Path(r"D:\BI文件\维度表\2026年维度表\Dim-仓库与物料对照表-2026年整理版.xlsx"),
    "dim-store-name": Path(r"D:\BI文件\维度表\2026年维度表\Dim-店铺名称汇总（金蝶&领星&简称）2026年整理版.xlsx"),
    "dim-customer-material": Path(r"D:\BI文件\维度表\2026年维度表\Dim-客户与物料对照表-2026年整理版.xlsx"),
    "dim-purchase-division": Path(r"D:\BI文件\2026年供应商名录\采购部分工明细表-2026.04.16.xlsx"),
}

SLOTS = {
    "fact-inventory": {"type": "fact", "title": "关账后库存事实表", "expectedName": "财务同步的表", "sheetHint": "", "skipRows": 0},
    "dim-product": {"type": "dimension", "title": "商品分类维表", "expectedName": "Dim-YL医疗器械商品分类-2026年整理版", "sheetHint": "Dim-YL医疗器械商品分类", "skipRows": 0},
    "dim-warehouse": {"type": "dimension", "title": "仓库、金蝶、旺店通、领星", "expectedName": "Dim-仓库_金蝶、旺店通、领星-2026年整理版", "sheetHint": "Dim-仓库汇总整理", "skipRows": 0},
    "dim-warehouse-material": {"type": "dimension", "title": "仓库物料事业部对照表", "expectedName": "Dim-仓库与物料对照表-2026年整理版", "sheetHint": "", "skipRows": 0},
    "dim-store-name": {"type": "dimension", "title": "维度 4", "expectedName": "维度 4", "sheetHint": "", "skipRows": 0},
    "dim-customer-material": {"type": "dimension", "title": "维度 5", "expectedName": "维度 5", "sheetHint": "", "skipRows": 0},
    "dim-purchase-division": {"type": "dimension", "title": "维度 6", "expectedName": "维度 6", "sheetHint": "", "skipRows": 0},
}


def norm(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def json_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return ""
        return value
    return str(value).strip()


def unique_headers(headers):
    counts = {}
    result = []
    for index, header in enumerate(headers):
        base = norm(header) or f"__EMPTY{'' if index == 0 else '_' + str(index)}"
        if base in counts:
            counts[base] += 1
            result.append(f"{base}_{counts[base]}")
        else:
            counts[base] = 0
            result.append(base)
    return result


def pick_sheet(workbook, hint):
    if hint:
        for name in workbook.sheetnames:
            if hint in name:
                return name
    return workbook.sheetnames[0]


def parse_file(slot_id, path):
    slot = SLOTS[slot_id]
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = pick_sheet(workbook, slot["sheetHint"])
    sheet = workbook[sheet_name]
    header_row = (slot.get("skipRows") or 0) + 1
    rows_iter = sheet.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows_iter, None)
    headers = unique_headers(next(rows_iter))
    rows = []
    for values in rows_iter:
        if not values or not any(value not in (None, "") for value in values):
            continue
        row = {}
        for header, value in zip(headers, values):
            row[header] = json_value(value)
        rows.append(row)
    workbook.close()

    stat = path.stat()
    saved_at = datetime.fromtimestamp(stat.st_mtime, timezone.utc).isoformat().replace("+00:00", "Z")
    record = {
        "id": slot_id,
        "type": slot["type"],
        "title": slot["title"],
        "expectedName": slot["expectedName"],
        "fileName": path.name,
        "size": stat.st_size,
        "lastModified": int(stat.st_mtime * 1000),
        "savedAt": saved_at,
        "appliedAt": saved_at,
        "sheetName": sheet_name,
        "rows": rows,
    }
    return record, headers


def normalize_header(value):
    return norm(value).replace("（", "(").replace("）", ")").replace(" ", "").lower()


def to_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = norm(value).replace(",", "").replace("￥", "").replace("¥", "").replace("元", "").replace(" ", "")
    try:
        return float(text) if text else 0.0
    except ValueError:
        return 0.0


def first_value(row, names):
    wanted = [normalize_header(name) for name in names]
    for key, value in row.items():
        if normalize_header(key) in wanted and norm(value) != "":
            return value
    return ""


def first_header_includes(row, words):
    wanted = [normalize_header(word) for word in words]
    for key, value in row.items():
        header = normalize_header(key)
        if all(word in header for word in wanted) and norm(value) != "":
            return value
    return ""


def ending_qty(row):
    candidates = [
        first_value(row, ["(结存)数量（库存）", "(结存)数量(库存)", "结存数量（库存）", "结存数量"]),
        first_header_includes(row, ["结存", "数量"]),
    ]
    values = list(row.values())
    if len(values) >= 7:
        candidates.append(values[6])
    for candidate in candidates:
        value = to_number(candidate)
        if value != 0 or norm(candidate) == "0":
            return value
    return 0.0


def material_code(row):
    return norm(first_value(row, ["物料编码"]))


def main():
    records = {}
    headers = {}
    for slot_id, path in FILES.items():
        record, file_headers = parse_file(slot_id, path)
        records[slot_id] = record
        headers[slot_id] = file_headers[:20]

    fact_rows = records["fact-inventory"]["rows"]
    summary = {
        "records": {
            key: {"fileName": value["fileName"], "rows": len(value["rows"]), "sheetName": value["sheetName"]}
            for key, value in records.items()
        },
        "factHeaderFirst20": headers["fact-inventory"],
        "factEndingQtyTotal": sum(ending_qty(row) for row in fact_rows if material_code(row)),
        "factNonZeroQtyRows": sum(1 for row in fact_rows if material_code(row) and ending_qty(row) != 0),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    payload = {
        "schemaVersion": 1,
        "project": "kcfx",
        "savedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "records": records,
    }
    out = Path(r"D:\BI文件\kcfx\data\shared-library.json")
    out.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"wrote {out} {out.stat().st_size}")


if __name__ == "__main__":
    main()
